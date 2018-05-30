from openpyxl import load_workbook
from openpyxl import Workbook

keyword = []
result = []
wbresult = Workbook()
sheet = wbresult.active
sheet.title = "test"
#关键字获取
wbkeyword = load_workbook("C:\\Users\\jinqiang\\Desktop\\key.xlsx")
for k in wbkeyword["Sheet1"]["A"]:
	keyword.append(k.value)
listresult = []
#数据集获取
wb = load_workbook("C:\\Users\\jinqiang\\Desktop\\test.xlsx")
for v in wb["Sheet1"]["C"]:
	listresult.append(v.value)
i = 1
for key in keyword:
	result = []
	result = list(filter(lambda x: key in x, listresult))
	sheet.cell(row = i, column = 1).value = key
	for n in range(2, len(result) + 2):
		sheet.cell(row = i, column = n).value = result[n - 2]
	i = i + 1
#保存结果excel
wbresult.save("C:\\Users\\jinqiang\\Desktop\\result.xlsx")