from openpyxl import load_workbook

wb = load_workbook("D:/highpython/Python-master/learn/01.xlsx")
print(wb.sheetnames)
sheet = wb.get_sheet_by_name("jin")
print(sheet["C"])
print(sheet["4"])
print(sheet["C4"].value)
print(sheet.max_row)
print(sheet.max_column)
for i in sheet["C"]:
    print(i.value, end=" ")